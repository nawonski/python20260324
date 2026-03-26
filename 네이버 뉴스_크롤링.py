import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# 네이버 검색 URL
url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query=%EB%B0%98%EB%8F%84%EC%B2%B4&ackey=u0vzgwgo"

# 페이지 요청
response = requests.get(url)
response.raise_for_status()  # 요청 실패 시 예외 발생

# BeautifulSoup으로 HTML 파싱
soup = BeautifulSoup(response.text, 'html.parser')

# 제공된 태그 구조를 기반으로 제목 추출 (YouTube 비디오 제목)
titles = []
headline_spans = soup.find_all('span', class_=lambda x: x and 'sds-comps-text-type-headline1' in x)
for span in headline_spans:
    title = span.get_text().strip()
    if title:  # 빈 제목 제외
        titles.append(title)

# openpyxl로 Excel 파일 생성 및 저장
wb = Workbook()
ws = wb.active
ws.title = "Crawled Titles"

# 헤더 추가
ws['A1'] = '제목'

# 제목들 추가
for idx, title in enumerate(titles, start=2):
    ws[f'A{idx}'] = title

# 파일 저장
wb.save('naver_result.xlsx')

print("크롤링 결과가 naver_result.xlsx 파일에 저장되었습니다.")