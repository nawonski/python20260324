import openpyxl
import random

# 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ProductList"

# 헤더 추가
ws['A1'] = '제품ID'
ws['B1'] = '제품명'
ws['C1'] = '가격'
ws['D1'] = '수량'

# 전자제품 이름 리스트
product_names = [
    '스마트폰', '노트북', '태블릿', '헤드폰', '스피커', '모니터', '키보드', '마우스',
    '충전기', '케이블', '카메라', '드론', '게임콘솔', 'TV', '냉장고', '세탁기',
    '에어컨', '청소기', '믹서기', '토스터', '커피머신', '전자레인지', '블렌더',
    '전기포트', '헤어드라이어', '면도기', '전동칫솔', '무선이어폰', '스마트워치',
    '피트니스 트래커', '프로젝터', '라우터', '외장하드', 'SSD', '메모리카드',
    '파워뱅크', '스마트홈 허브', '보안카메라', '도어락', '로봇청소기', '공기청정기',
    '가습기', '제습기', '전기난로', '선풍기', '에어프라이어', '오븐', '식기세척기',
    '세탁기', '건조기'
]

# 데이터 생성 및 추가
for i in range(2, 102):  # 2부터 101까지 (100개)
    product_id = f"P{i-1:03d}"  # P001, P002, ...
    product_name = random.choice(product_names)
    price = random.randint(10000, 1000000)  # 10,000원 ~ 1,000,000원
    quantity = random.randint(1, 100)  # 1 ~ 100개

    ws[f'A{i}'] = product_id
    ws[f'B{i}'] = product_name
    ws[f'C{i}'] = price
    ws[f'D{i}'] = quantity

# 파일 저장
wb.save('ProductList.xlsx')
print("ProductList.xlsx 파일이 생성되었습니다.")
